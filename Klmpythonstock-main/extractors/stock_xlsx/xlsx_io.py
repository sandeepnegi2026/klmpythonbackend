import io
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.header_match import match_header

from extractors.stock_xlsx.parse_common import cell_text


def workbook_kind(file_bytes, filename=""):
    if file_bytes[:2] == b"PK":
        return ".xlsx"
    if file_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return ".xls"
    head = file_bytes[:4096].lower()
    if file_bytes[:3] == b"\xef\xbb\xbf" or b"<table" in head or b"<html" in head:
        return ".html"
    return Path(filename).suffix.lower() if filename else ".xlsx"


def read_sheets(file_bytes, filename=""):
    ext = workbook_kind(file_bytes, filename)
    engines = ["openpyxl", "xlrd"] if ext == ".xlsx" else ["xlrd", "openpyxl"]
    last_error = None
    for engine in engines:
        try:
            return pd.ExcelFile(io.BytesIO(file_bytes), engine=engine), engine
        except Exception as exc:
            last_error = exc
    raise last_error or ValueError("Could not read workbook")


def unmerge_xlsx(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    for ws in workbook.worksheets:
        for merged_range in list(ws.merged_cells.ranges):
            value = ws.cell(merged_range.min_row, merged_range.min_col).value
            coord = str(merged_range)
            ws.unmerge_cells(coord)
            for row in ws.iter_rows(
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
            ):
                for cell in row:
                    cell.value = value
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def sheet_rows(df):
    rows = []
    for raw in df.itertuples(index=False, name=None):
        row = [cell_text(value) for value in raw]
        while row and row[-1] == "":
            row.pop()
        if any(row):
            rows.append(row)
    return rows


def sheet_score(rows):
    # Only cells containing a letter can match an (alphabetic) header synonym; numeric,
    # date and empty cells always score 0. Skipping them before the expensive fuzzy
    # match_header keeps the score identical while cutting ~90% of the calls on a
    # data-heavy sheet (the header scan was the dominant extraction cost).
    score = 0
    for row in rows[:150]:
        for cell in row:
            if any(ch.isalpha() for ch in cell) and match_header(cell, "stock")[0]:
                score += 1
    return score


def load_rows(file_bytes, filename, requested_sheet=None):
    kind = workbook_kind(file_bytes, filename)
    if kind == ".xlsx":
        try:
            file_bytes = unmerge_xlsx(file_bytes)
        except Exception:
            pass
    xls, _ = read_sheets(file_bytes, filename)
    candidates = (
        [requested_sheet]
        if requested_sheet and requested_sheet in xls.sheet_names
        else xls.sheet_names
    )
    # With a single candidate there is nothing to compare, so sheet_score (the
    # dominant cost on data-heavy books) would be computed and then ignored. Skip
    # straight to that sheet's rows — identical result, no wasted header scan.
    if len(candidates) == 1:
        return candidates[0], sheet_rows(xls.parse(candidates[0], header=None))
    best = (None, [], -1)
    for sheet in candidates:
        rows = sheet_rows(xls.parse(sheet, header=None))
        score = sheet_score(rows)
        if requested_sheet == sheet or score > best[2]:
            best = (sheet, rows, score)
        if requested_sheet == sheet:
            break
    return best[0], best[1]
