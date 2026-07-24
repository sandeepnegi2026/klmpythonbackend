import io
import json
import os
import traceback

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from core.canonical import CANONICAL_FIELDS
from core.header_match import match_header, normalize, set_header_overrides
from core.io_helpers import build_csv, build_json, build_xlsx, rows_to_dataframe
from core.quality import build_quality
from core.scoring import coverage
from core.triage import _to_float
from extractors import party_pdf, party_xlsx, stock_pdf, stock_xlsx

HEADER_MATCH_THRESHOLD = 0.62  # mirrors core.header_match.match_header default

st.set_page_config(page_title="Pharma Report Extractor", page_icon="PDF", layout="wide")

MAX_UPLOAD_BYTES = 25 * 1024 * 1024
REPORT_TYPES = {
    "Party-wise Sales": "party",
    "Stock & Sales Report": "stock",
}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
FORMAT_EXTENSIONS = {
    "PDF": ".pdf",
    "Excel (.xlsx / .xls)": ".xlsx",
}
ROUTES = {
    ("party", ".pdf"): party_pdf.extract,
    ("party", ".xlsx"): party_xlsx.extract,
    ("stock", ".pdf"): stock_pdf.extract,
    ("stock", ".xlsx"): stock_xlsx.extract,
}


def _xlsx_sheet_names(file_bytes, filename):
    ext = os.path.splitext(filename.lower())[1]
    if ext == ".xls":
        try:
            import pandas as pd

            return pd.ExcelFile(io.BytesIO(file_bytes), engine="xlrd").sheet_names
        except Exception:
            return []
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        return workbook.sheetnames
    except Exception:
        return []


def _detected_layout(result):
    debug = result.get("debug") or {}
    label = debug.get("layout_label") or debug.get("format_label") or "Unknown"
    key = debug.get("layout") or debug.get("detected_format") or ""
    return label, key


def _coverage_chip(chip):
    icons = {"ok": "OK", "warn": "WARN", "missing": "MISSING"}
    source = chip.get("source_header") or "not matched"
    st.markdown(f"**{icons.get(chip['status'], 'MISSING')} `{chip['field']}`**")
    st.caption(f"source: {source}")


def _render_warnings(warnings):
    """Two kinds of warnings share result['warnings']; frame them honestly.

    Reconciliation/sanity warnings are a vendor SOURCE-FILE issue — the extraction is
    faithful, the vendor's own numbers just don't add up. Structural warnings ("no rows",
    "no text / scanned", header/parse errors) are a real extraction/file problem and must
    NOT be labelled 'faithful'. Sanity warnings are authored to contain the token
    'source-file' (see stock_{pdf,xlsx}/postprocess.py) — that's the split signal."""
    warnings = warnings or []
    if not warnings:
        return
    sanity = [w for w in warnings if "source-file" in w]
    structural = [w for w in warnings if "source-file" not in w]
    if structural:
        st.warning("\n".join(f"- {w}" for w in structural))
    if sanity:
        st.caption(
            "⚠️ Data check — the vendor's own numbers don't fully add up. The extraction is "
            "faithful to the file; please spot-check these products against the original report."
        )
        st.warning("\n".join(f"- {w}" for w in sanity))


def _render_coverage(result, report_type):
    chips = coverage(result, report_type)
    if not chips:
        st.info("No required fields configured.")
        return
    cols = st.columns(min(4, len(chips)))
    for idx, chip in enumerate(chips):
        with cols[idx % len(cols)]:
            _coverage_chip(chip)
    _render_warnings(result.get("warnings"))


def _render_rows(result, uploaded_name, report_type):
    rows = result.get("rows", []) or []
    if not rows:
        st.info("No structured rows extracted.")
        return
    df = rows_to_dataframe(rows, report_type)
    st.dataframe(df, use_container_width=True, hide_index=True, height=520)
    stem = os.path.splitext(uploaded_name)[0]
    c1, c2, c3 = st.columns(3)
    c1.download_button("Download CSV", data=build_csv(rows, report_type), file_name=f"{stem}_extracted.csv", mime="text/csv")
    c2.download_button(
        "Download XLSX",
        data=build_xlsx(rows, report_type),
        file_name=f"{stem}_extracted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    c3.download_button("Download JSON", data=build_json(result), file_name=f"{stem}_result.json", mime="application/json")


def _render_layout(result):
    label, key = _detected_layout(result)
    st.markdown(f"**Detected layout:** {label}")
    if key:
        st.caption(f"Internal key: `{key}`")
    pages = result.get("pages", []) or []
    if not pages:
        st.info("No per-page metadata captured.")
        return
    st.dataframe(pd.DataFrame(pages), use_container_width=True, hide_index=True)


def _best_header_row(rows, report_type, max_scan=150):
    """Best-effort: return the cells of the row most likely to be the header.

    Independent of whether the layout parser succeeded — it scans the raw cells and
    picks the row with the most DISTINCT canonical matches, then returns ALL of that
    row's cells (including the ones that did NOT match). That is what lets us show an
    unmatched ``party_name`` column even when header detection failed entirely.
    """
    best_cells, best_n = [], 0
    for row in rows[:max_scan]:
        cells = ["" if c is None else str(c) for c in row]
        keys = {match_header(c, report_type)[0] for c in cells}
        keys.discard(None)
        if len(keys) > best_n:
            best_cells, best_n = cells, len(keys)
    return best_cells


def _raw_headers(result, report_type):
    """The file's actual column headers, recovered for diagnosis.

    1. xlsx previews store ``raw_text`` as tab-joined rows — reconstruct and find the
       header row (works even when extraction produced 0 rows).
    2. Otherwise fall back to the source headers the extractor already recorded.
    """
    raw_text = result.get("raw_text") or ""
    if "\t" in raw_text:
        recon = [line.split("\t") for line in raw_text.splitlines()[:150]]
        cells = [c for c in _best_header_row(recon, report_type) if str(c).strip()]
        if cells:
            return cells
    return [str(h) for h in (result.get("headers_detected") or {})]


def _render_verdict(triage):
    # 4-state display: the wire bucket stays GREEN/AMBER/RED, but AMBER splits on
    # extraction_ok — True means the numbers are proven correct (they match the
    # report's own printed totals) and only the vendor's data doesn't add up;
    # None means correctness could not be auto-confirmed (YELLOW, needs a check).
    bucket = triage.get("bucket", "?")
    code = triage.get("reason_code", "")
    reason = triage.get("reason", "")
    ok = triage.get("extraction_ok")
    if bucket == "GREEN":
        st.success(f"🟢 **Correct**\n\n{reason}")
    elif bucket == "RED":
        st.error(f"🔴 **Not correct**\n\n{reason}")
    elif bucket == "AMBER" and (ok is True or (ok is None and str(reason).startswith("Extraction is correct"))):
        st.warning(f"🟠 **Correct — vendor data mismatch**\n\n{reason}")
    elif bucket == "AMBER":
        st.warning(f"🟡 **Needs a quick check**\n\n{reason}")
    else:
        st.warning(f"**{bucket}**\n\n{reason}")
    st.caption(f"technical: `{bucket} · {code}`")


# --------------------------------------------------------------------------- #
# Analytics & Statistics — an at-a-glance aggregate view of the extracted rows,
# mirroring the /admin/uploads InsightCard panel. Computed purely from the
# vendor's own as-printed values (PSUI has no price catalog), so a column the
# report never printed (commonly PTS) shows "—" rather than a fabricated 0.
# --------------------------------------------------------------------------- #
def _indian_group(digits):
    """Group an all-digit integer string Indian-style: last 3, then 2s. '5193088' -> '51,93,088'."""
    if len(digits) <= 3:
        return digits
    head, tail = digits[:-3], digits[-3:]
    parts = []
    while len(head) > 2:
        parts.insert(0, head[-2:])
        head = head[:-2]
    parts.insert(0, head)
    return ",".join(parts) + "," + tail


def _fmt_indian_number(value, decimals=2, strip=True):
    """Indian-grouped number, up to `decimals` places, no currency symbol.

    `strip=True` drops trailing zeros (mirrors JS maximumFractionDigits, e.g.
    23507 -> "23,507", 101.32 -> "101.32"); `strip=False` keeps exactly `decimals`
    places (used by the currency formatter). None -> em dash."""
    if value is None:
        return "—"
    try:
        n = float(value)
    except (TypeError, ValueError):
        return "—"
    neg = n < 0
    n = abs(n)
    if decimals > 0:
        whole = int(n)
        frac = round(n - whole, decimals)
        if frac >= 1:                       # rounding carried into the integer part
            whole += int(frac)
            frac -= int(frac)
        frac_str = f"{frac:.{decimals}f}"[2:]
        if strip:
            frac_str = frac_str.rstrip("0")
        out = _indian_group(str(whole)) + (f".{frac_str}" if frac_str else "")
    else:
        out = _indian_group(str(int(round(n))))
    return ("-" + out) if neg else out


def _fmt_inr(value, decimals=2):
    """Rupee amount, Indian grouping, ₹ prefix, always `decimals` places
    (mirrors formatINRPrice, e.g. ₹51,93,088.05). None -> em dash (no ₹)."""
    if value is None:
        return "—"
    formatted = _fmt_indian_number(value, decimals, strip=False)
    if formatted == "—":
        return "—"
    return ("-₹" + formatted[1:]) if formatted.startswith("-") else ("₹" + formatted)


def _is_return(row):
    # PSUI rows carry no transaction_type; a return is a negative qty or amount.
    return (_to_float(row.get("qty")) or 0) < 0 or (_to_float(row.get("amount")) or 0) < 0


def _distinct_count(rows, *fields):
    """Distinct non-empty count over the first populated field per row (trim+lower)."""
    seen = set()
    for row in rows:
        for field in fields:
            value = row.get(field)
            if value is not None and str(value).strip():
                seen.add(str(value).strip().lower())
                break
    return len(seen)


def _sum_col(rows, *fields, transform=None, require_nonzero=False):
    """Sum across rows. None if NO row carries a parseable value in any given field.

    `require_nonzero=True` also returns None when every parseable value is 0 — used
    for money fields, because enforce_schema fills a column the report never printed
    with "0", and a card should show "—" for that, not a misleading ₹0.00. Quantities
    leave it False so a genuine 0 (e.g. no free qty) still shows as 0."""
    total, seen, nonzero = 0.0, False, False
    for row in rows:
        value = None
        for field in fields:
            value = _to_float(row.get(field))
            if value is not None:
                break
        if value is None:
            continue
        seen = True
        if value != 0:
            nonzero = True
        total += transform(row, value) if transform else value
    if not seen or (require_nonzero and not nonzero):
        return None
    return total


def _analytics_row(label, value, emphasis=False, accent=None):
    return (label, value, emphasis, accent)


def _party_cards(rows):
    return_rows = [r for r in rows if _is_return(r)]

    products = _distinct_count(rows, "canonical_name", "product_name")
    parties = _distinct_count(rows, "normalized_party_name", "party_name")

    # Headline qty + money totals sum EVERY row (returns/credit lines included) so
    # the panel equals the report's own printed GRAND TOTAL and the exported TOTAL
    # row (core/io_helpers.append_totals_row, which sums the additive columns over
    # all rows). An extraction preview must reconcile to the vendor's printed
    # figures — a report prints ONE arithmetic column sum, it does not net returns
    # out of gross like the catalog-based admin tiles do. Return/Net below stay as
    # an informational split (Return Qty = Σ|return-row qty|; Net = Gross − Return).
    gross_qty = _sum_col(rows, "qty") or 0.0
    return_qty = _sum_col(return_rows, "qty", transform=lambda r, v: abs(v)) or 0.0
    net_qty = gross_qty - return_qty
    free_qty = _sum_col(rows, "free_qty") or 0.0

    # Financial totals come straight from what the report PRINTS (PSUI has no price
    # catalog — see the module note; the admin's MRP/PTR/PTS need catalog defaults we
    # don't have). require_nonzero collapses an all-"0" (never-printed) column to "—".
    mrp = _sum_col(rows, "mrp", transform=lambda r, v: (_to_float(r.get("qty")) or 0.0) * v, require_nonzero=True)
    sale_value = _sum_col(rows, "amount", require_nonzero=True)
    taxable = _sum_col(rows, "taxable_value", require_nonzero=True)
    gst_amount = _sum_col(rows, "gst_amount", require_nonzero=True)
    net_amount = _sum_col(rows, "net_amount", require_nonzero=True)

    avg_qty = (gross_qty / products) if products else 0.0

    def line_value(row):
        for field in ("amount", "taxable_value"):
            value = _to_float(row.get(field))
            if value is not None:
                return value
        qty, rate = _to_float(row.get("qty")), _to_float(row.get("rate"))
        return qty * rate if (qty is not None and rate is not None) else None

    line_vals = [lv for lv in (line_value(r) for r in rows) if lv is not None]
    total_value = sum(line_vals) if line_vals else None
    avg_value = (total_value / products) if (total_value is not None and products) else None

    return [
        ("Invoice Overview", [
            _analytics_row("Products", _fmt_indian_number(products, 0)),
            _analytics_row("Parties", _fmt_indian_number(parties, 0)),
            _analytics_row("Avg Qty / Product", _fmt_indian_number(avg_qty, 2), emphasis=True),
            _analytics_row("Avg Value / Product", _fmt_inr(avg_value, 2)),
        ]),
        ("Financial Summary", [
            _analytics_row("MRP", _fmt_inr(mrp, 2)),
            _analytics_row("Sale Value", _fmt_inr(sale_value, 2), emphasis=True),
            _analytics_row("Taxable Value", _fmt_inr(taxable, 2)),
            _analytics_row("GST Amount", _fmt_inr(gst_amount, 2)),
            _analytics_row("Net Amount", _fmt_inr(net_amount, 2)),
        ]),
        ("Quantity Summary", [
            _analytics_row("Gross Qty", _fmt_indian_number(gross_qty, 2)),
            _analytics_row("Return Qty", _fmt_indian_number(return_qty, 2), accent="warning"),
            _analytics_row("Net Qty", _fmt_indian_number(net_qty, 2), emphasis=True),
            _analytics_row("Free Qty", _fmt_indian_number(free_qty, 2)),
        ]),
    ]


def _stock_cards(rows):
    products = _distinct_count(rows, "canonical_name", "product_name")

    opening = _sum_col(rows, "opening_stock") or 0.0
    purchase_qty = _sum_col(rows, "purchase_stock") or 0.0
    sales_qty = _sum_col(rows, "sales_qty") or 0.0
    closing = _sum_col(rows, "closing_stock") or 0.0

    purchase_val = _sum_col(rows, "purchase_value", require_nonzero=True)
    sales_val = _sum_col(rows, "sales_value", require_nonzero=True)
    closing_val = _sum_col(rows, "closing_stock_value", require_nonzero=True)

    avg_purchase = (purchase_qty / products) if products else 0.0
    avg_sales = (sales_qty / products) if products else 0.0

    return [
        ("Stock Overview", [
            _analytics_row("Products", _fmt_indian_number(products, 0)),
            _analytics_row("Avg Purchase / Product", _fmt_indian_number(avg_purchase, 2)),
            _analytics_row("Avg Sales / Product", _fmt_indian_number(avg_sales, 2), emphasis=True),
        ]),
        ("Value Summary", [
            _analytics_row("Purchase Value", _fmt_inr(purchase_val, 2)),
            _analytics_row("Sales Value", _fmt_inr(sales_val, 2), emphasis=True),
            _analytics_row("Closing Value", _fmt_inr(closing_val, 2)),
        ]),
        ("Quantity Summary", [
            _analytics_row("Opening Stock", _fmt_indian_number(opening, 2)),
            _analytics_row("Purchase Qty", _fmt_indian_number(purchase_qty, 2)),
            _analytics_row("Sales Qty", _fmt_indian_number(sales_qty, 2), emphasis=True),
            _analytics_row("Closing Stock", _fmt_indian_number(closing, 2)),
        ]),
    ]


def _render_insight_card(title, rows):
    with st.container(border=True):
        st.markdown(
            f"<div style='font-size:0.72rem;font-weight:700;letter-spacing:0.06em;"
            f"text-transform:uppercase;color:#64748b;margin-bottom:0.4rem;'>{title}</div>",
            unsafe_allow_html=True,
        )
        for label, value, emphasis, accent in rows:
            if accent == "warning":
                color = "#b45309"
            elif emphasis:
                color = "#2563eb"
            else:
                color = "inherit"
            size, weight = ("1.35rem", "700") if emphasis else ("0.95rem", "600")
            st.markdown(
                "<div style='display:flex;justify-content:space-between;align-items:baseline;padding:0.12rem 0;'>"
                f"<span style='font-size:0.8rem;color:#475569;'>{label}</span>"
                f"<span style='font-size:{size};font-weight:{weight};color:{color};text-align:right;'>{value}</span>"
                "</div>",
                unsafe_allow_html=True,
            )


def _render_analytics(result, report_type):
    rows = result.get("rows", []) or []
    if not rows:
        return
    st.divider()
    st.markdown("### Analytics & Statistics")
    cards = _party_cards(rows) if report_type == "party" else _stock_cards(rows)
    cols = st.columns(len(cards))
    for col, (title, card_rows) in zip(cols, cards):
        with col:
            _render_insight_card(title, card_rows)


def _render_header_mapping(result, report_type, checks, raw_headers):
    if not raw_headers:
        st.info("No header row could be recovered for this file (it may be parsed positionally).")
        return

    active = st.session_state.get("overrides") or {}
    active_norm = {normalize(h): f for h, f in active.items()}

    table, unmatched = [], []
    for header in raw_headers:
        nh = normalize(header)
        if nh in active_norm:
            key, score, method = active_norm[nh], 1.0, "override"
        else:
            key, score, method = match_header(header, report_type, min_score=0.0)
        matched = method == "override" or score >= HEADER_MATCH_THRESHOLD
        table.append({
            "header": header,
            "mapped_to": key or "—",
            "score": round(score, 3),
            "needs": HEADER_MATCH_THRESHOLD,
            "status": "override" if method == "override" else ("yes" if matched else "NO"),
        })
        if not matched:
            unmatched.append({"header": header, "closest_field": key or "—", "score": round(score, 3)})

    data_missing = checks.get("data_missing") or []
    if data_missing:
        st.error(
            "Required field(s) not detected: "
            + ", ".join(f"`{f}`" for f in data_missing)
            + f". The likely cause is an unmatched column below — it scored under the "
            f"{HEADER_MATCH_THRESHOLD} match threshold, so its data was dropped. "
            "Map it to the right field below and re-run."
        )

    st.markdown("**Every header in this file and what it matched:**")
    st.dataframe(pd.DataFrame(table), use_container_width=True, hide_index=True)

    if active:
        st.divider()
        st.success("Active manual overrides: " + ", ".join(f"`{h}` → `{f}`" for h, f in active.items()))
        fields_const = "PARTY_FIELDS" if report_type == "party" else "STOCK_FIELDS"
        st.markdown(
            "**Make it permanent** — add these synonyms in `core/canonical.py`, then mirror to "
            "`Backends/` and run regression (AGENTS.md):"
        )
        st.code(
            "\n".join(
                f'# add "{str(h).lower().strip()}" to {fields_const}["{f}"]["synonyms"]'
                for h, f in active.items()
            ),
            language="python",
        )

    # Editable set = unmatched headers + headers currently force-mapped by an override.
    # An applied override makes its header "matched" (so it is not in `unmatched`); add
    # it back here so it can be revised or set back to "(leave unmapped)" individually,
    # not only wiped via "Clear all overrides".
    editable = list(unmatched)
    unmatched_norms = {normalize(u["header"]) for u in unmatched}
    for header, field in active.items():
        if normalize(header) not in unmatched_norms:
            editable.append({"header": header, "closest_field": field, "score": 1.0, "active": True})

    if editable:
        st.divider()
        if unmatched:
            st.markdown("**Unmatched headers — data from these columns is currently dropped:**")
            st.dataframe(pd.DataFrame(unmatched), use_container_width=True, hide_index=True)
        st.markdown("**Fix a column live:** map a header to a field (or back to *(leave unmapped)*), then re-run.")
        field_options = ["(leave unmapped)"] + list(CANONICAL_FIELDS[report_type].keys())
        with st.form("override_form"):
            for i, item in enumerate(editable):
                header = item["header"]
                # Two physical columns can share a normalized header (e.g. the doubled
                # "CrQty" in the KLM Venus banded layout), which would collide on one
                # widget key — Streamlit rejects duplicate keys. The list index makes
                # each key unique; the readback below uses this exact stored key.
                wkey = f"ov::{report_type}::{i}::{normalize(header)}"
                item["_wkey"] = wkey
                if wkey not in st.session_state:
                    st.session_state[wkey] = active.get(header, "(leave unmapped)")
                if item.get("active"):
                    hint = f" — currently overridden → `{item['closest_field']}`"
                elif item["closest_field"] != "—":
                    hint = f" — closest: `{item['closest_field']}` ({item['score']})"
                else:
                    hint = ""
                st.selectbox(f"`{header}`{hint}", field_options, key=wkey)
            col_a, col_b = st.columns(2)
            apply_btn = col_a.form_submit_button("Apply overrides & re-run", type="primary")
            clear_btn = col_b.form_submit_button("Clear all overrides")
        if apply_btn:
            # Rebuild from every editable row, so clearing one (set to unmapped) drops
            # just that override while the rest are preserved.
            new_overrides = {}
            for item in editable:
                header = item["header"]
                choice = st.session_state.get(item["_wkey"], "(leave unmapped)")
                if choice != "(leave unmapped)":
                    new_overrides[header] = choice
            st.session_state["overrides"] = new_overrides
            st.rerun()
        if clear_btn:
            for k in [k for k in list(st.session_state.keys()) if str(k).startswith(f"ov::{report_type}::")]:
                del st.session_state[k]
            st.session_state["overrides"] = {}
            st.rerun()
    else:
        st.success("Every recovered header is mapped to a canonical field.")

    detected = result.get("headers_detected") or {}
    if detected:
        st.divider()
        st.markdown("**Final mapping the extractor used (source header → canonical key):**")
        st.dataframe(
            pd.DataFrame([{"source_header": s, "canonical_key": k} for s, k in detected.items()]),
            use_container_width=True,
            hide_index=True,
        )


def _render_triage(quality):
    triage = quality["triage"]
    checks = quality["checks"]
    _render_verdict(triage)
    st.divider()

    cols = st.columns(4)
    cols[0].metric("Score %", f"{round(quality['score_pct'] * 100)}%")
    cols[1].metric("Rows", checks.get("row_count", 0))
    mm = checks.get("product_master_match_rate")
    cols[2].metric("Master match", "n/a" if mm is None else f"{round(mm * 100)}%")
    eff = (checks.get("sanity") or {}).get("effective_pass_rate")
    cols[3].metric("Stock sanity", "n/a" if eff is None else f"{round(eff * 100)}%")

    if checks.get("data_missing"):
        st.error("Hard-missing data fields (blocking): " + ", ".join(checks["data_missing"]))
    if checks.get("soft_missing"):
        st.warning("Soft-missing fields (layout-dependent, not blocking): " + ", ".join(checks["soft_missing"]))

    st.divider()
    st.markdown("**All cross-checks:**")
    st.json(checks)


def _settings_sidebar(uploaded, report_type_key, ext):
    settings = {}
    with st.sidebar:
        if ext == ".pdf":
            settings["strategy"] = st.radio("PDF strategy", ["lattice", "stream", "hybrid"], index=0 if report_type_key == "stock" else 2)
            settings["page_range"] = st.text_input("Page range", value="", placeholder="all, or e.g. 1-3,5")
            c1, c2 = st.columns(2)
            settings["x_tolerance"] = c1.number_input("X tolerance", min_value=0.5, max_value=20.0, value=3.0, step=0.5)
            settings["y_tolerance"] = c2.number_input("Y tolerance", min_value=0.5, max_value=20.0, value=3.0, step=0.5)
            settings["snap_tolerance"] = st.number_input("Snap tolerance", min_value=0.5, max_value=20.0, value=3.0, step=0.5)
        elif uploaded is not None:
            sheet_names = _xlsx_sheet_names(uploaded.getvalue(), uploaded.name)
            sheet_options = ["Auto"] + sheet_names
            selected_sheet = st.selectbox("Sheet", sheet_options, index=0)
            settings["sheet_name"] = None if selected_sheet == "Auto" else selected_sheet
            header_mode = st.radio("Header row hint", ["Auto", "Row index"], index=0)
            settings["header_row"] = None
            if header_mode == "Row index":
                settings["header_row"] = st.number_input("Header row index", min_value=1, value=1, step=1)
        settings["manual_row_count"] = st.number_input("Manual row count for score", min_value=0, value=0, step=1)
    return settings


st.title("Pharma Report Extractor Spike")
st.caption("Four-route shell: report type x file format, with shared coverage and scoring.")

with st.sidebar:
    st.header("Input")
    report_label = st.radio("Report Type", list(REPORT_TYPES.keys()), index=0)
    format_label = st.radio("File Format", list(FORMAT_EXTENSIONS.keys()), index=0)
    is_pdf = format_label == "PDF"
    ext = FORMAT_EXTENSIONS[format_label]
    upload_types = ["pdf"] if is_pdf else ["xlsx", "xls"]
    uploaded = st.file_uploader(f"Upload {format_label}", type=upload_types, help="Max 25 MB")

report_type_key = REPORT_TYPES[report_label]
settings = _settings_sidebar(uploaded, report_type_key, ext)

with st.sidebar:
    run_btn = st.button("Run", use_container_width=True, type="primary")

if uploaded is None:
    st.info("Choose report type, file format, upload a file, then click Run.")
    st.stop()

if uploaded.size > MAX_UPLOAD_BYTES:
    st.error(f"File too large ({uploaded.size / 1024 / 1024:.1f} MB). Max is 25 MB.")
    st.stop()

actual_ext = os.path.splitext(uploaded.name.lower())[1]
if is_pdf and actual_ext != ".pdf":
    st.error(f"Selected {format_label}, but uploaded file extension is `{actual_ext}`.")
    st.stop()
if not is_pdf and actual_ext not in EXCEL_EXTENSIONS:
    st.error(f"Selected {format_label}, but uploaded file extension is `{actual_ext}`.")
    st.stop()

# Manual header overrides are per-file: reset them (and their widgets) when the
# uploaded file or report type changes, so a stale mapping never leaks across files.
ov_sig = (uploaded.name, uploaded.size, report_type_key)
if st.session_state.get("overrides_sig") != ov_sig:
    for k in [k for k in list(st.session_state.keys()) if str(k).startswith("ov::")]:
        del st.session_state[k]
    st.session_state["overrides"] = {}
    st.session_state["overrides_sig"] = ov_sig
overrides = st.session_state.get("overrides") or {}

cache_key = (
    uploaded.name,
    uploaded.size,
    report_type_key,
    ext,
    json.dumps(settings, sort_keys=True, default=str),
    json.dumps(overrides, sort_keys=True),
)
if run_btn or st.session_state.get("cache_key") != cache_key:
    try:
        with st.spinner("Extracting..."):
            route_ext = ".pdf" if is_pdf else ".xlsx"
            fn = ROUTES[(report_type_key, route_ext)]
            settings["filename"] = uploaded.name
            set_header_overrides(report_type_key, overrides)
            try:
                result = fn(uploaded.getvalue(), settings)
            finally:
                set_header_overrides(report_type_key, None)
        result["file_name"] = uploaded.name
        result["report_type"] = report_type_key
        result["file_format"] = ext
        st.session_state["result"] = result
        st.session_state["cache_key"] = cache_key
    except Exception:
        st.error(f"Extraction failed:\n\n```\n{traceback.format_exc()}\n```")
        st.stop()

result = st.session_state.get("result")
if not result:
    st.stop()

quality = build_quality(result, report_type_key, settings.get("manual_row_count") or None)
triage = quality["triage"]
checks = quality["checks"]
raw_headers = _raw_headers(result, report_type_key)  # recovered once; reused by both tabs
layout_label, layout_key = _detected_layout(result)

summary_cols = st.columns(5)
summary_cols[0].metric("Rows", len(result.get("rows", []) or []))
summary_cols[1].metric("Headers", sum(1 for value in (result.get("headers_detected") or {}).values() if value))
summary_cols[2].metric("Score", f"{quality['score']}/10")
summary_cols[3].metric("Runtime", f"{result.get('elapsed_ms', 0)} ms")
summary_cols[4].metric("Detected Layout", layout_label)
if layout_key:
    st.caption(f"Layout key: `{layout_key}`")

_render_verdict(triage)

_render_analytics(result, report_type_key)

row_tab_label = "Stock Rows" if report_type_key == "stock" else "Extracted Rows"
(
    tab_coverage,
    tab_headers,
    tab_rows,
    tab_raw,
    tab_layout,
    tab_triage,
    tab_warnings,
    tab_json,
    tab_score,
) = st.tabs([
    "Coverage",
    "Header Mapping",
    row_tab_label,
    "Raw Text / Sheet Preview",
    "Layout Debug",
    "Triage / Checks",
    "Sanity Warnings",
    "JSON Dump",
    "Score",
])

with tab_coverage:
    _render_coverage(result, report_type_key)
    detected = result.get("headers_detected") or {}
    if detected:
        st.divider()
        st.dataframe(
            pd.DataFrame([{"source_header": source, "canonical_key": target} for source, target in detected.items()]),
            use_container_width=True,
            hide_index=True,
        )

with tab_headers:
    _render_header_mapping(result, report_type_key, checks, raw_headers)

with tab_rows:
    _render_rows(result, uploaded.name, report_type_key)

with tab_raw:
    st.code(result.get("raw_text") or "(no raw text or sheet preview captured)", language=None)

with tab_layout:
    _render_layout(result)

with tab_triage:
    _render_triage(quality)

with tab_warnings:
    warnings = result.get("warnings", []) or []
    sanity = result.get("sanity", {}) or {}
    if sanity:
        st.json(sanity)
    if warnings:
        _render_warnings(warnings)
    else:
        st.success("No warnings.")

with tab_json:
    diagnostic = {
        "file_name": result.get("file_name"),
        "report_type": result.get("report_type"),
        "detected_layout": layout_label,
        "triage": triage,
        "checks": checks,
        "score": quality["score"],
        "score_pct": quality["score_pct"],
        "parts": quality["parts"],
        "header_mapping": raw_headers,
        "headers_detected": result.get("headers_detected"),
        "warnings": result.get("warnings"),
        "debug": result.get("debug"),
        "rows_sample": (result.get("rows") or [])[:20],
    }
    st.markdown("**Copy this block and paste it to me to fix this file:**")
    st.code(json.dumps(diagnostic, indent=2, default=str), language="json")
    st.download_button(
        "Download full diagnostic JSON",
        data=json.dumps({"diagnostic": diagnostic, "result": result}, indent=2, default=str),
        file_name=f"{os.path.splitext(uploaded.name)[0]}_diagnostic.json",
        mime="application/json",
    )
    st.divider()
    st.markdown("**Full raw result:**")
    st.json(result)

with tab_score:
    st.metric("Upload score", f"{quality['score']}/10")
    st.dataframe(
        pd.DataFrame([{"check": key, "points": value} for key, value in quality["parts"].items()]),
        use_container_width=True,
        hide_index=True,
    )
    st.caption("Row-count points require the optional manual row count in the sidebar.")
