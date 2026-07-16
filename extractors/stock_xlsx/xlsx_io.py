import io
from pathlib import Path
from statistics import median

import pandas as pd
from openpyxl import load_workbook

from core.header_match import match_header

from extractors.stock_xlsx.parse_common import cell_text

# A tab is kept as a real data sheet when its score is at least this fraction of the MEDIAN
# tab score. Division-per-tab books (COSMO/DERMA/...) all score in one band so the whole
# cluster passes (median, unlike max, is not skewed by one outlier-high tab); a stray
# blank/summary tab scores far below the median and is dropped.
KEEP_RATIO = 0.6


def workbook_kind(file_bytes, filename=""):
    if file_bytes[:2] == b"PK":
        return ".xlsx"
    if file_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return ".xls"
    head = file_bytes[:4096].lower()
    if file_bytes[:3] == b"\xef\xbb\xbf" or b"<table" in head or b"<html" in head:
        return ".html"
    return Path(filename).suffix.lower() if filename else ".xlsx"


# Excel-2003 "XML Spreadsheet" exports ship as plain SpreadsheetML text with a .xls
# extension; both pandas engines reject XML, so this path previously raised and the
# file triaged as SCANNED_OR_EMPTY with 0 rows. The byte signature below (XML
# declaration + Excel.Sheet progid) never occurs in a real .xls/.xlsx/HTML export,
# so previously-readable workbooks never enter the conversion branch.
_SPREADSHEETML_MARK = b'mso-application progid="Excel.Sheet"'


def _spreadsheetml_to_xlsx(file_bytes):
    """Convert SpreadsheetML Worksheet/Table/Row/Cell into real .xlsx bytes."""
    import xml.etree.ElementTree as ET

    from openpyxl import Workbook

    ns = "{urn:schemas-microsoft-com:office:spreadsheet}"
    root = ET.fromstring(file_bytes)
    book = Workbook()
    book.remove(book.active)
    for ws_el in root.findall(f"{ns}Worksheet"):
        title = (ws_el.get(f"{ns}Name") or f"Sheet{len(book.sheetnames) + 1}")[:31]
        sheet = book.create_sheet(title=title)
        table = ws_el.find(f"{ns}Table")
        if table is None:
            continue
        r = 0
        for row_el in table.findall(f"{ns}Row"):
            r = int(row_el.get(f"{ns}Index", r + 1))
            c = 0
            for cell_el in row_el.findall(f"{ns}Cell"):
                c = int(cell_el.get(f"{ns}Index", c + 1))
                span = int(cell_el.get(f"{ns}MergeAcross", 0))
                data = cell_el.find(f"{ns}Data")
                text = "" if data is None else "".join(data.itertext())
                if text:
                    # replicate across the merge span, like unmerge_xlsx does
                    for col in range(c, c + span + 1):
                        sheet.cell(row=r, column=col, value=text)
                c += span
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def read_sheets(file_bytes, filename=""):
    if file_bytes[:256].lstrip().startswith(b"<?xml") and _SPREADSHEETML_MARK in file_bytes[:512]:
        file_bytes = _spreadsheetml_to_xlsx(file_bytes)
    ext = workbook_kind(file_bytes, filename)
    engines = ["openpyxl", "xlrd"] if ext == ".xlsx" else ["xlrd", "openpyxl"]
    last_error = None
    for engine in engines:
        try:
            return pd.ExcelFile(io.BytesIO(file_bytes), engine=engine), engine
        except Exception as exc:
            last_error = exc
    raise last_error or ValueError("Could not read workbook")


def unmerge_xlsx(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    for ws in workbook.worksheets:
        for merged_range in list(ws.merged_cells.ranges):
            value = ws.cell(merged_range.min_row, merged_range.min_col).value
            coord = str(merged_range)
            ws.unmerge_cells(coord)
            for row in ws.iter_rows(
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
            ):
                for cell in row:
                    cell.value = value
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def sheet_rows(df):
    rows = []
    for raw in df.itertuples(index=False, name=None):
        row = [cell_text(value) for value in raw]
        while row and row[-1] == "":
            row.pop()
        if any(row):
            rows.append(row)
    return rows


def sheet_score(rows):
    # Only cells containing a letter can match an (alphabetic) header synonym; numeric,
    # date and empty cells always score 0. Skipping them before the expensive fuzzy
    # match_header keeps the score identical while cutting ~90% of the calls on a
    # data-heavy sheet (the header scan was the dominant extraction cost).
    score = 0
    for row in rows[:150]:
        for cell in row:
            if any(ch.isalpha() for ch in cell) and match_header(cell, "stock")[0]:
                score += 1
    return score


def load_rows(file_bytes, filename, requested_sheet=None):
    kind = workbook_kind(file_bytes, filename)
    if kind == ".xlsx":
        try:
            file_bytes = unmerge_xlsx(file_bytes)
        except Exception:
            pass
    xls, _ = read_sheets(file_bytes, filename)
    candidates = (
        [requested_sheet]
        if requested_sheet and requested_sheet in xls.sheet_names
        else xls.sheet_names
    )
    # With a single candidate there is nothing to compare, so sheet_score (the
    # dominant cost on data-heavy books) would be computed and then ignored. Skip
    # straight to that sheet's rows — identical result, no wasted header scan.
    if len(candidates) == 1:
        return candidates[0], sheet_rows(xls.parse(candidates[0], header=None))
    best = (None, [], -1)
    for sheet in candidates:
        rows = sheet_rows(xls.parse(sheet, header=None))
        score = sheet_score(rows)
        if requested_sheet == sheet or score > best[2]:
            best = (sheet, rows, score)
        if requested_sheet == sheet:
            break
    return best[0], best[1]


def load_data_sheets(file_bytes, filename, requested_sheet=None):
    """Return ``[(sheet_name, rows), ...]`` for every genuine data tab of the workbook.

    A multi-tab workbook (e.g. one tab per division) must have ALL its data tabs
    extracted, not just the single best-scoring tab that ``load_rows`` returns. The same
    ``sheet_score`` heuristic tells real report tabs from blank/summary tabs:
      * an explicit ``requested_sheet``  -> just that tab (UI dropdown override);
      * <= 1 non-empty tab, or nothing scores -> the single best tab (``load_rows`` parity,
        so unknown formats never accidentally multi-merge);
      * otherwise every tab scoring at least ``KEEP_RATIO`` of the MEDIAN tab score, in
        workbook order (median, not max, so one outlier-high tab can't evict the cluster).
    """
    kind = workbook_kind(file_bytes, filename)
    if kind == ".xlsx":
        try:
            file_bytes = unmerge_xlsx(file_bytes)
        except Exception:
            pass
    xls, _ = read_sheets(file_bytes, filename)
    if requested_sheet and requested_sheet in xls.sheet_names:
        rows = sheet_rows(xls.parse(requested_sheet, header=None))
        return [(requested_sheet, rows)] if rows else []
    scored = []
    for sheet in xls.sheet_names:
        rows = sheet_rows(xls.parse(sheet, header=None))
        if not rows:
            continue
        scored.append((sheet, rows, sheet_score(rows)))
    if not scored:
        return []
    best = max(score for _, _, score in scored)
    if best <= 0 or len(scored) == 1:
        # Can't distinguish data from noise (or only one tab): keep the single best, which
        # max() picks as the first among ties -- exactly load_rows()'s strict-greater choice.
        top = max(scored, key=lambda item: item[2])
        return [(top[0], top[1])]
    ref = median([score for _, _, score in scored if score > 0])
    keep = [(sheet, rows) for sheet, rows, score in scored if score > 0 and score >= ref * KEEP_RATIO]
    return keep or [(scored[0][0], scored[0][1])]
